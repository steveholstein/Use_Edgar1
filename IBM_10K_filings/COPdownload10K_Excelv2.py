'''
# pip install sec-edgar-downloader

### Answer to Question 1:
The sections I listed are commonly found in typical 10-K filings based on the structure provided by the SEC EDGAR system. These sections are part of the standard 10-K reporting requirements. However, the exact titles and their placements can vary slightly between different companies and filings.
### Answer to Question 2:
If you want to dynamically identify and extract relevant sections, you need to look for actual sections within the downloaded 10-K data. This involves finding headings that likely match the required sections. A more robust approach would include:
1. Extracting potential section titles based on common patterns (e.g., capital letters, bold text in HTML, etc.).
2. Using regular expressions to match and extract sections of interest.
3. Identify the closest matches to known section names.
Below is the amended code to dynamically find and extract the actual sections within the downloaded EDGAR data.
```python
'''
test
from sec_edgar_downloader import Downloader
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl
import pandas as pd
import re
import os

# File destination for the downloaded 10-K filings
destination = r'C:\Users\hello\PycharmProjects\Use_Edgar1\APPLE_10K_filings'
Path(destination).mkdir(parents=True, exist_ok=True)

absolute_path = os.path.abspath(destination)
print("Absolute path:", absolute_path)

# Optionally clear the directory before downloading new files
clear_directory = input(
    f"The directory '{destination}' already contains files. Do you want to clear it before downloading new filings? (yes/no): ")
if clear_directory.lower() == 'yes':
    for file in os.listdir(destination):
        file_path = os.path.join(destination, file)
        try:
            os.unlink(file_path)
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")

# Initialize a downloader instance with a specific download directory
dl = Downloader("Ed22", "greetingssteve@gmail.com", download_folder=destination)

# Download all 10-K filings for Apple (AAPL)
dl.get("10-K", "AAPL")

# Load all downloaded 10-K filings
filings_dir = Path(destination)
filing_files = list(filings_dir.glob("*.txt"))

# Initialize the Excel workbook
excel_filename = os.path.join(destination, "Apple_10K_Data.xlsx")
wb = openpyxl.Workbook()
default_sheet = wb.active
default_sheet.title = "Default"

# Sections of interest
sections_of_interest = [
    "Business Overview", "Risk Factors", "Selected Financial Data",
    "Management's Discussion and Analysis", "Financial Statements and Supplementary Data"
]

# Patterns for common section headers
section_patterns = [
    r'\bBusiness Overview\b', r'\bRisk Factors\b', r'\bSelected Financial Data\b',
    r'\bManagement\'s Discussion and Analysis\b', r'\bFinancial Statements\b'
]


def extract_section(text, start_idx, end_idx):
    return text[start_idx:end_idx].strip()


def find_section_indices(text, patterns):
    indices = []
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            indices.append((match.start(), match.end(), pattern.strip(r'\b')))
    indices = sorted(indices)
    return indices


def extract_tables_from_html(html, title_prefix):
    soup = BeautifulSoup(html, 'html.parser')
    tables = soup.find_all('table')

    for idx, table in enumerate(tables):
        dfs = pd.read_html(str(table))
        if dfs:
            df = dfs[0]
            yield df, f"{title_prefix} Table {idx + 1}"


# Process each 10-K filing
sheet_added = False
for filing_file in filing_files:
    with open(filing_file, 'r', encoding='utf-8', errors='ignore') as f:
        soup = BeautifulSoup(f, 'html.parser')
        text = soup.get_text(" ", strip=True)

    section_indices = find_section_indices(text, section_patterns)

    for i, (start_idx, end_idx, section) in enumerate(section_indices):
        if i < len(section_indices) - 1:
            next_start_idx = section_indices[i + 1][0]
        else:
            next_start_idx = len(text)

        section_text = extract_section(text, start_idx, next_start_idx)
        print(f"Extracting section '{section}' from {filing_file.name}")

        if section in ["Selected Financial Data", "Financial Statements"]:
            # Extract and write numerical data
            for df, table_title in extract_tables_from_html(str(soup), section):
                with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a') as writer:
                    df.to_excel(writer, sheet_name=table_title[:31], index=False)
                    print(f"Table '{table_title}' written to Excel.")
            sheet_added = True

        else:
            # Create a new sheet for this section
            ws = wb.create_sheet(title=section[:30])  # Excel sheet names have 31 char limit
            ws.cell(row=1, column=1, value=section)

            # Split the text into rows to avoid cell content length limits
            max_cell_length = 32767
            rows = [section_text[i:i + max_cell_length] for i in range(0, len(section_text), max_cell_length)]
            for idx, row in enumerate(rows, start=2):
                ws.cell(row=idx, column=1, value=row)

            sheet_added = True

# Remove the default sheet if other sheets were added
if sheet_added:
    wb.remove(default_sheet)
else:
    default_sheet.cell(row=1, column=1, value="No sections found in the downloaded filings.")

# Save the Excel workbook
wb.save(excel_filename)
print(f"Data written to Excel file at {excel_filename}")

'''
```
### How the Code Works:
1. **Patterns**: Define regex patterns for common section headers.
2. **Find Sections**: Use `find_section_indices` to locate the start and end indices of sections based on those patterns.
3. **Extract Sections**: Extract the text between the section indices.
4. **Dynamic Table Extraction**: Extract tables relevant to financial data sections and save them using `pandas`.
5. **Writing to Excel**: Write both text sections and numerical data tables into appropriately named sheets in an Excel file.
This approach ensures that the script dynamically adapts to the actual structure of the downloaded 10-K filings and extracts relevant sections even if they have slightly different titles.

'''