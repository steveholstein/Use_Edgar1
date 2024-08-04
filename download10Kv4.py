
from sec_edgar_downloader import Downloader
from pathlib import Path
from bs4 import BeautifulSoup
import openpyxl
import os

# File destination for the downloaded 10-K filings
destination = r'C:\Users\hello\PycharmProjects\Use_Edgar1\APPLE_10K_filings'
Path(destination).mkdir(parents=True, exist_ok=True)

absolute_path = os.path.abspath(destination)
print("Absolute path:", absolute_path)

# Initialize a downloader instance with a specific download directory
dl = Downloader("Ed22", "greetingssteve@gmail.com", download_folder=destination)

# Download all 10-K filings for Apple (AAPL)
dl.get("10-K", "AAPL")

# Load all downloaded 10-K filings
filings_dir = Path(destination)
filing_files = list(filings_dir.glob("*.txt"))

# Initialize the Excel workbook
excel_filename = os.path.join(destination, "Apple_10K_Data.xlsx")
wb = openpyxl.Workbook()
default_sheet = wb.active
default_sheet.title = "Default"

# Sections of interest
sections_of_interest = [
    "Business Overview", "Risk Factors", "Selected Financial Data",
    "Management's Discussion and Analysis", "Financial Statements and Supplementary Data"
]

def extract_section(text, section_title):
    start_idx = text.find(section_title)
    if start_idx == -1:
        return ""
    end_idx = len(text)
    for next_section in sections_of_interest:
        if next_section == section_title:
            continue
        next_idx = text.find(next_section, start_idx + len(section_title))
        if start_idx < next_idx < end_idx:
            end_idx = next_idx
    return text[start_idx:end_idx]

# Process each 10-K filing
sheet_added = False
for filing_file in filing_files:
    with open(filing_file, 'r', encoding='utf-8', errors='ignore') as f:
        soup = BeautifulSoup(f, 'html.parser')
        text = soup.get_text(" ", strip=True)

    for section in sections_of_interest:
        section_text = extract_section(text, section)

        if section_text:
            # Create a new sheet for this section
            ws = wb.create_sheet(title=section[:30])  # Excel sheet names have 31 char limit
            ws.cell(row=1, column=1, value=section)
            ws.cell(row=2, column=1, value=section_text)
            sheet_added = True

            if section == "Selected Financial Data":
                # Here you can parse numerical data from the section_text
                # and add it to the appropriate tab.
                pass

# Remove the default sheet if other sheets were added
if sheet_added:
    wb.remove(default_sheet)

# Save the Excel workbook
wb.save(excel_filename)
print(f"Data written to Excel file at {excel_filename}")


'''
### Details and Notes:
1. **BeautifulSoup**: I used BeautifulSoup for parsing HTML content from the filings.
2. **OpenPyXL**: This library is used to handle the Excel writing. Each section gets its own sheet.
3. **Section Extraction Function**: The `extract_section` function takes care of finding the text between section headers.
4. **Excel File Handling**: An Excel workbook is initialized, and sheets are added based on sections found in the filings.
5. **Financial Data Parsing**: You should add specific logic to parse the numerical data from the "Selected Financial Data" section and any other sections you are interested in.

Make sure you've installed the required Python packages:
```sh
pip install sec-edgar-downloader beautifulsoup4 openpyxl pandas
```

This script is a starting point, and you might need additional parsing logic for identifying and handling numerical data in the financial sections accurately.
'''
